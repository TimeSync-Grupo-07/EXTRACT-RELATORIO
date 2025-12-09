import os
import time
import mysql.connector
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl.cell.cell
from flask import Flask, send_file, jsonify
from flask_cors import CORS
import io

app = Flask(__name__)
CORS(app)  # Permite requisições do frontend

def gerar_relatorio():
    # ---------------- CONFIGURAÇÃO via ENV ---------------- #
    DB_HOST = os.getenv("DB_HOST", "localhost")
    DB_PORT = int(os.getenv("DB_PORT", 3306))
    DB_USER = os.getenv("DB_USER", "teste")
    DB_PASS = os.getenv("DB_PASS", "urubu100")
    DB_NAME = os.getenv("DB_NAME", "Timesync")

    db_config = {
        "host": DB_HOST,
        "port": DB_PORT,
        "user": DB_USER,
        "password": DB_PASS,
        "database": DB_NAME
    }

    views = {
        "Relatório de Projeto": "vw_relatorio_projeto",
        "Alocação de Recursos": "vw_alocacao_recursos",
        "Indicadores de Eficiência": "vw_indicadores_eficiencia",
        "Comparativo Mensal": "vw_comparativo_mensal"
    }

    # ---------------- TENTATIVA DE CONEXÃO COM RETRY ---------------- #
    max_retries = 12
    retry_delay = 5

    conn = None
    for attempt in range(1, max_retries + 1):
        try:
            conn = mysql.connector.connect(**db_config)
            print(f"Conectado ao MySQL em {DB_HOST} (tentativa {attempt})")
            break
        except Exception as e:
            print(f"Falha conexão (tentativa {attempt}/{max_retries}): {e}")
            if attempt == max_retries:
                raise
            time.sleep(retry_delay)

    # ---------------- ESTILOS ---------------- #
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    section_font = Font(color="FFFFFF", bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ---------------- CRIAÇÃO DO EXCEL ---------------- #
    wb = Workbook()
    wb.remove(wb.active)

    # Cria as abas separadas
    for aba_nome, view_nome in views.items():
        ws = wb.create_sheet(title=aba_nome)
        df = pd.read_sql(f"SELECT * FROM {view_nome}", conn)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

        # Ajusta largura das colunas
        for i, column_cells in enumerate(ws.columns, 1):
            valid_cells = [cell for cell in column_cells if not isinstance(cell, openpyxl.cell.cell.MergedCell)]
            if not valid_cells:
                continue
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in valid_cells)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2
            for cell in valid_cells:
                cell.border = border

    # ---------------- ABA "RESUMO COMPLETO" ---------------- #
    ws_all = wb.create_sheet(title="Resumo Completo")

    # Obtém dados das views
    data_relatorio = pd.read_sql("SELECT * FROM vw_relatorio_projeto", conn)
    data_recursos = pd.read_sql("SELECT * FROM vw_alocacao_recursos", conn)
    data_eficiencia = pd.read_sql("SELECT * FROM vw_indicadores_eficiencia", conn)
    data_comparativo = pd.read_sql("SELECT * FROM vw_comparativo_mensal", conn)

    # Nome do projeto e mês atual
    if not data_relatorio.empty and 'nome_projeto' in data_relatorio.columns:
        # Pega o primeiro projeto ou concatena todos
        projetos = data_relatorio["nome_projeto"].unique()
        if len(projetos) == 1:
            projeto_nome = projetos[0]
        else:
            projeto_nome = f"{len(projetos)} Projetos"
    else:
        projeto_nome = "PROJETO"
    
    mes_atual = datetime.now().strftime("%B %Y").upper()

    # ---------------- TÍTULO ---------------- #
    ws_all.merge_cells("A1:K1")
    title_cell = ws_all["A1"]
    title_cell.value = f"RELATÓRIO MENSAL DO PROJETO {projeto_nome.upper()} - {mes_atual}"
    title_cell.fill = header_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = center_align

    row_pos = 3

    # ---------------- FUNÇÃO DE SEÇÃO ---------------- #
    def add_section(title, df, row_pos):
        ws_all.merge_cells(start_row=row_pos, start_column=1, end_row=row_pos, end_column=11)
        section_cell = ws_all.cell(row=row_pos, column=1, value=title)
        section_cell.fill = header_fill
        section_cell.font = section_font
        section_cell.alignment = center_align
        row_pos += 1

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), row_pos):
            ws_all.append(row)
            if r_idx == row_pos:
                for cell in ws_all[r_idx]:
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = center_align
            for cell in ws_all[r_idx]:
                cell.border = border
        row_pos += len(df) + 2
        return row_pos

    # Adiciona seções
    row_pos = add_section("ALOCAÇÃO DE RECURSOS", data_recursos, row_pos)
    row_pos = add_section("INDICADORES DE EFICIÊNCIA (PWC)", data_eficiencia, row_pos)
    row_pos = add_section("COMPARATIVO MENSAL", data_comparativo, row_pos)

    # ---------------- OBSERVAÇÕES AUTOMÁTICAS ---------------- #
    ws_all.merge_cells(start_row=row_pos, start_column=1, end_row=row_pos, end_column=11)
    section_cell = ws_all.cell(row=row_pos, column=1, value="OBSERVAÇÕES AUTOMÁTICAS")
    section_cell.fill = header_fill
    section_cell.font = section_font
    section_cell.alignment = center_align
    row_pos += 1

    obs = [
        ("Custo real > estimado em 10%", "O projeto está acima do custo previsto para o mês em questão."),
        ("Taxa de erro > 5%", "Existem inconsistências nos apontamentos. Avaliar colaboradores com maior índice de erro."),
        ("Aderência < 85%", "O cronograma de horas está abaixo do esperado. Risco de atraso detectado.")
    ]

    ws_all.append(["Situação", "Mensagem"])
    for cell in ws_all[row_pos]:
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = center_align

    for o in obs:
        ws_all.append(o)
        for cell in ws_all[ws_all.max_row]:
            cell.border = border

    # Ajusta largura das colunas
    for i, column_cells in enumerate(ws_all.columns, 1):
        valid_cells = [cell for cell in column_cells if not isinstance(cell, openpyxl.cell.cell.MergedCell)]
        if not valid_cells:
            continue
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in valid_cells)
        ws_all.column_dimensions[get_column_letter(i)].width = max_length + 2

    conn.close()

    # Salva em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

@app.route('/api/exportar-relatorio', methods=['GET'])
def exportar_relatorio():
    try:
        excel_file = gerar_relatorio()
        nome_arquivo = f"Relatorio_Timesync_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        # Adicione logging detalhado
        import traceback
        error_details = traceback.format_exc()
        print(f"ERRO DETALHADO: {error_details}")
        return jsonify({
            "error": "Falha ao gerar relatório",
            "details": str(e),
            "traceback": error_details
        }), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({"status": "healthy", "service": "relatorio-export"})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)